import csv
import openpyxl
from datetime import datetime

username_id_dict = {}
user_dict_activity = {}

def generate_white_list():
    workbook = openpyxl.load_workbook("white_list.xlsx")
    sheet = workbook.active
    white_list = []

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        white_list.append(int(row[0]))

    return white_list

def entrance_or_exit(location):
    ENTRANCE_LIST = ["AccessEngine.Devices.Kołowrót 4 - wej.Zdarzenie", "AccessEngine.Devices.Kołowrót 3 - wej.Zdarzenie"]
    EXIT_LIST = ["AccessEngine.Devices.Kołowrót 6 - wyj.Zdarzenie", "AccessEngine.Devices.Kołowrót 5 - wyj.Zdarzenie"]
    if location in ENTRANCE_LIST:
        return "entrance"
    elif location in EXIT_LIST:
        return "exit"

def convert_to_datetime(datetime_str):
    datetime_str = datetime_str.rsplit(" ", 1)[0]
    date_format = '%Y-%m-%d %H:%M:%S'
    date_object = datetime.strptime(datetime_str, date_format)
    return date_object

def parse_activity_row(row, title_dict):
    activity_time = convert_to_datetime(row[title_dict["time"]])
    full_name = f"{row[title_dict['firstname']]} {row[title_dict['name']]}"
    user_id = int(row[title_dict["id"]])
    entrance_exit = entrance_or_exit(row[title_dict["location"]])
    return activity_time, full_name, user_id, entrance_exit

def initialize_user(user_id, full_name):
    user_dict_activity[user_id] = {"full_name": full_name}
    user_dict_activity[user_id]["activity"] = []
    user_dict_activity[user_id]["work_schedule"] = []


def get_user_activity_dict(worker_to_check):
    """
    user_dict_activity = { "id": {"full_name":"str", work_schedule:[(date_start, start_hour,finish_hour)...], "activity":[(date_time, entrance_or_exit)...]} ...}
    """
    WHITE_LIST = generate_white_list()
    with open("15-19.csv", "r", encoding="utf-8") as activity_file:
        csv_file = csv.reader(activity_file)
        activity_list = list(csv_file)

        title = activity_list[0][0].split(";")
        title_dict = {
            "time": title.index("Czas zdarzenia"),
            "firstname": title.index("FIRSTNAME"),
            "name": title.index("NAME"),
            "id": title.index("CARDNO"),
            "location": title.index("Adres")
        }

        for row in activity_list[1:]:
            row = row[0].split(";")
            activity_time, full_name, user_id, entrance_exit = parse_activity_row(row, title_dict)

            if user_id in WHITE_LIST:
                continue

            if user_id not in worker_to_check:
                continue

            if full_name not in username_id_dict:
                username_id_dict[full_name] = user_id

            if user_id not in user_dict_activity:
                initialize_user(user_id, full_name)

            user_dict_activity[user_id]["activity"].append((activity_time, entrance_exit))

    with open("odmowa dostępu.csv", "r", encoding="utf-8") as refusal_file:
        csv_refusal_file = csv.reader(refusal_file)
        refusal_list = list(csv_refusal_file)

        title_refusal = refusal_list[0][0].split(";")
        title_refusal_dict = {
            "time": title_refusal.index("Czas zdarzenia"),
            "firstname": title_refusal.index("FIRSTNAME"),
            "name": title_refusal.index("NAME")
        }

        for row in refusal_list[1:]:
            row = row[0].split(";")
            time = convert_to_datetime(row[title_refusal_dict["time"]])
            firstname = row[title_refusal_dict["firstname"]]
            name = row[title_refusal_dict["name"]]

            full_name = f"{firstname} {name}"
            worker_id = username_id_dict.get(full_name, None)
            if not worker_id:
                continue

            user_dict_activity[worker_id]["activity"].append((time, "exit"))
            user_dict_activity[worker_id]["activity"] = sorted(user_dict_activity[worker_id]["activity"], key=lambda x: x[0])


    return user_dict_activity
