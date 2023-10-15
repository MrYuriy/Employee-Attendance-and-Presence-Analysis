import csv
import openpyxl
from datetime import datetime, time, timedelta, date
import xlwt


WORKER_ID_DICT = {}
WORKER_DICT_ACTIVITY = {}
WHITE_LIST = {}
MIN_MAX_RANGE_HOUR_TO_CHEK = 6
CSV_DIAPAZONE = ()
RAPORT_LIST = []

WHITE_LIST_PATH = "white_list.xlsx"
SCHEDULE_PATH = "Grafik Testowy.xlsx"
SCHEDULE_MONTH = "Wrzesień"
ACTIVITY_FILE_PATH = "15-19.csv"
REFUSAL_FILE_PATH = "odmowa dostępu.csv"


def generate_white_list():
    workbook = openpyxl.load_workbook(WHITE_LIST_PATH)
    sheet = workbook.active
    white_list = []

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        white_list.append(int(row[0]))

    return white_list

def entrance_or_exit(location):
    ENTRANCE_LIST = ["AccessEngine.Devices.Kołowrót 4 - wej.Zdarzenie", "AccessEngine.Devices.Kołowrót 3 - wej.Zdarzenie"]
    EXIT_LIST = ["AccessEngine.Devices.Kołowrót 6 - wyj.Zdarzenie", "AccessEngine.Devices.Kołowrót 5 - wyj.Zdarzenie"]
    if location in ENTRANCE_LIST:
        return "entrance"
    elif location in EXIT_LIST:
        return "exit"

def convert_to_datetime(datetime_str):
    datetime_str = datetime_str.rsplit(" ", 1)[0]
    date_format = '%Y-%m-%d %H:%M:%S'
    date_object = datetime.strptime(datetime_str, date_format)
    return date_object

def parse_activity_row(row, title_dict):
    activity_time = convert_to_datetime(row[title_dict["time"]])
    full_name = f"{row[title_dict['firstname']]} {row[title_dict['name']]}"
    card_id = int(row[title_dict["id"]])
    entrance_exit = entrance_or_exit(row[title_dict["location"]])
    return activity_time, full_name, card_id, entrance_exit


def generate_workers_schedule_and_information():
    workbook = openpyxl.load_workbook(SCHEDULE_PATH)
    worksheet = workbook[SCHEDULE_MONTH]
    title_list = [cell.value for cell in worksheet[2]]
    title_dict = {
        "id": title_list.index("Nr karty"),
        "department": title_list.index("Dział"),
        "firm": title_list.index("Firma"),
        "position": title_list.index("Stanowisko"),
        "boss": title_list.index("Przełożony"),
        "full_name": title_list.index("Nazwisko i imię")

    }

    start_schedule_index = next((i for i, item in enumerate(title_list) if isinstance(item, datetime)), None)

    for row in worksheet.iter_rows(min_row=3, values_only=True):
        work_schedule = []
        card_id = row[title_dict["id"]]
        department = row[title_dict["department"]]
        firm = row[title_dict["firm"]]
        position = row[title_dict["position"]]
        boss = row[title_dict["boss"]]
        full_name = row[title_dict["full_name"]]

        if card_id not in WORKER_DICT_ACTIVITY:
            WORKER_DICT_ACTIVITY[card_id] = {}
        row = row[start_schedule_index:]
        for index, value in enumerate(row[::2]):
            work_schedule.append((title_list[index*2+start_schedule_index] ,value, row[index*2+1]))

        WORKER_DICT_ACTIVITY[card_id]["card_id"] = card_id
        WORKER_DICT_ACTIVITY[card_id]["department"] = department
        WORKER_DICT_ACTIVITY[card_id]["firm"] = firm
        WORKER_DICT_ACTIVITY[card_id]["position"] = position
        WORKER_DICT_ACTIVITY[card_id]["full_name"] = full_name
        WORKER_DICT_ACTIVITY[card_id]["boss"] = boss
        WORKER_DICT_ACTIVITY[card_id]["work_schedule"] = work_schedule
        WORKER_DICT_ACTIVITY[card_id]["activity"] = []
        


def generate_worker_activity():
    global WHITE_LIST
    WHITE_LIST = generate_white_list()

    with open(ACTIVITY_FILE_PATH, "r", encoding="utf-8") as activity_file:
        csv_file = csv.reader(activity_file)
        activity_list = list(csv_file)

        title = activity_list[0][0].split(";")
        title_dict = {
            "time": title.index("Czas zdarzenia"),
            "firstname": title.index("FIRSTNAME"),
            "name": title.index("NAME"),
            "id": title.index("CARDNO"),
            "location": title.index("Adres")
        }



        global CSV_DIAPAZONE 
        CSV_DIAPAZONE= (
            convert_to_datetime(activity_list[-1][0].split(";")[0]),
            convert_to_datetime(activity_list[1][0].split(";")[0]) + timedelta(hours=8)
            )

        for row in activity_list[1:]:
            row = row[0].split(";")
            activity_time, full_name, card_id, entrance_exit = parse_activity_row(row, title_dict)

            if card_id in WHITE_LIST or card_id not in WORKER_DICT_ACTIVITY:
                continue

            if full_name not in WORKER_ID_DICT:
                WORKER_ID_DICT[full_name] = card_id
            WORKER_DICT_ACTIVITY[card_id]["activity"].append((activity_time, entrance_exit))

    with open(REFUSAL_FILE_PATH, "r", encoding="utf-8") as refusal_file:
        csv_refusal_file = csv.reader(refusal_file)
        refusal_list = list(csv_refusal_file)

        title_refusal = refusal_list[0][0].split(";")
        title_refusal_dict = {
            "time": title_refusal.index("Czas zdarzenia"),
            "firstname": title_refusal.index("FIRSTNAME"),
            "last name": title_refusal.index("NAME")
        }

        for row in refusal_list[1:]:
            row = row[0].split(";")
            time = convert_to_datetime(row[title_refusal_dict["time"]])
            if CSV_DIAPAZONE[0] <= time <= CSV_DIAPAZONE[1]:
                firstname = row[title_refusal_dict["firstname"]]
                lastname = row[title_refusal_dict["last name"]]
                full_name = f"{firstname} {lastname}"
                worker_id = WORKER_ID_DICT.get(full_name, None)
                if not worker_id:
                    continue

                WORKER_DICT_ACTIVITY[worker_id]["activity"].append((time, "exit"))
                WORKER_DICT_ACTIVITY[worker_id]["activity"] = sorted(WORKER_DICT_ACTIVITY[worker_id]["activity"], key=lambda x: x[0])


def get_date_finish_shif(start_hour, finish_hour, start_day):
    if finish_hour < start_hour:
        return datetime.combine(start_day, finish_hour) + timedelta(days=1)
    return datetime.combine(start_day, finish_hour)


def get_actual_start_time(activity_list, shadule_start_time):
    """ return actuale datetime going to work if worker late return status 'yes' """
    filtered_times = [time for time in activity_list if time <= shadule_start_time]
    status = None

    if filtered_times:
        actual_start_time = max(filtered_times, key=lambda x: x)
        return actual_start_time, status

    else:
        actual_start_time = min(activity_list, key=lambda x: x)
        if actual_start_time.day == shadule_start_time.day:
            status = "yes"
            return actual_start_time, status
        else:
            #print("брак даних")
            return None, status


def get_actual_finish_time(activity_list, shadule_finish_time):
    """ return actuale datetime going out work if worker erlier return status 'yes' """
    filtered_times = [time for time in activity_list if time >= shadule_finish_time]
    status = None

    if filtered_times:
        actual_finish_time = min(filtered_times, key=lambda x: x)
        return actual_finish_time, status

    else:
        actual_finish_time = max(activity_list, key=lambda x: x)
        if actual_finish_time.day == shadule_finish_time.day:
            status = "yes"
            return actual_finish_time, status
        else:
            #print("брак даних")
            return None, status


def get_total_breakfast(user_activity):
    total_time = timedelta()
    user_activity = user_activity[1:-1]
    exit = None

    if user_activity:
        for datetime_, action in user_activity:
            if action == "exit":
                exit = datetime_
            elif action == "entrance" and exit is not None:
                total_time += datetime_ - exit
    return total_time


def analize_breakfast(start_datetime, finish_datetime, length_lunch):
    time_difference = finish_datetime - start_datetime
    work_hours = time_difference.total_seconds() / 3600

    if work_hours <= 8 and length_lunch > 30:
        return length_lunch
    if work_hours > 8 and length_lunch > 60:
        return length_lunch
    return ""

def get_work_total_hours( 
        datetime_start_shift, 
        datetime_finish_shift, 
        datetime_start_actualy,
        datetime_finish_actualy,
        breacfest_total):
    start_datetime = datetime_start_shift
    finsh_datetime = datetime_finish_shift
    if datetime_start_actualy and datetime_start_actualy > datetime_start_shift:
        start_datetime = datetime_start_actualy
    if datetime_finish_actualy and datetime_finish_actualy < datetime_finish_shift:
        finsh_datetime = datetime_finish_actualy
    return str(finsh_datetime - start_datetime - breacfest_total)


def analize_schedule():
    for worker in WORKER_DICT_ACTIVITY:
        worker_information = WORKER_DICT_ACTIVITY[worker]
    #worker_information = WORKER_DICT_ACTIVITY[2705366509]
        schedule_list_information = worker_information["work_schedule"]
        activity = sorted(worker_information["activity"], key=lambda x: x[0])

        for shift in schedule_list_information:
            date_start_shift, start_hour,finish_hour = shift
            total_breacfest = time(0,0,0)

            #якщо тип змінної start_hour не дата пропустити ітерацію
            if not isinstance(start_hour, time):
                continue

                # початок кінець робочої зміни
            datetime_finish_shift = get_date_finish_shif(start_hour, finish_hour, date_start_shift)
            datetime_start_shift = datetime.combine(date_start_shift, start_hour)

            #діапазон в якому потрібно витягнути пересування працівника
            diapazone_start_work = datetime_start_shift - timedelta(hours=MIN_MAX_RANGE_HOUR_TO_CHEK)
            diapazone_finish_work = datetime_finish_shift + timedelta(hours=MIN_MAX_RANGE_HOUR_TO_CHEK)

            PRINT = False

            work_period_activity = [] #work_period_activity = [(datetime_action, action)...]
            work_mix_activity = [] #work_mix_activity = datetime_action]
            enterences_list = []
            exites_list = []

            # remark_row = [
            #     date_start_shift.strftime("%Y-%m-%d"), 
            #     worker_information["department"],
            #     worker_information["full_name"],
            #     worker_information["card_id"],
            #     worker_information["firm"], 
            #     worker_information["position"], 
            #     worker_information["boss"], 
            #     "", "", "", "", ""
            # ]

            remark_row_dict = {
                "date_start_shift": date_start_shift.strftime("%Y-%m-%d"), 
                "hour_start_shif":str(start_hour),
                "hour_finish_shif":str(finish_hour),
                "hour_start_work": "",
                "hour_finish_work": "",
                "department": worker_information["department"],
                "full_name": worker_information["full_name"],
                "card_id": worker_information["card_id"],
                "firm": worker_information["firm"], 
                "position": worker_information["position"], 
                "boss": worker_information["boss"], 
                "breakfest_total_time": "",
                "being_late": "",
                "early_exit": "",
                "work_total_time": "",
                "notes": ""
            }

            # створення списку дійі часу працівника
            for datetime_action, action in  activity:
                if diapazone_start_work <= datetime_action<= diapazone_finish_work:
                    if action == "entrance":
                        enterences_list.append(datetime_action)
                        work_period_activity.append((datetime_action, action))
                    if action == "exit":
                        exites_list.append(datetime_action)
                        work_period_activity.append((datetime_action, action))

            work_mix_activity = enterences_list + exites_list

            if enterences_list:
                actual_start_time, status = get_actual_start_time(enterences_list, datetime_start_shift)
                if actual_start_time:
                    remark_row_dict["hour_start_work"] = actual_start_time.strftime('%Y-%m-%d %H:%M')
                    for index, time_action in enumerate(work_mix_activity):
                        if time_action == actual_start_time:
                            work_mix_activity = work_mix_activity[index:]
                if status:
                    PRINT = True
                    #remark_row[8] = actual_start_time.strftime('%Y-%m-%d %H:%M')
                    remark_row_dict["being_late"] = actual_start_time.strftime('%Y-%m-%d %H:%M')

            if exites_list:
                actual_finish_time, status = get_actual_finish_time(exites_list, datetime_finish_shift)
                if actual_finish_time:
                    remark_row_dict["hour_finish_work"] = actual_finish_time.strftime('%Y-%m-%d %H:%M')
                    for index, time_action in enumerate(work_mix_activity):
                        if time_action == actual_finish_time:
                            work_mix_activity = work_mix_activity[:index+1]
                if status:
                    PRINT = True
                    #remark_row[9] = actual_finish_time.strftime('%Y-%m-%d %H:%M')
                    remark_row_dict["early_exit"] = actual_finish_time.strftime('%Y-%m-%d %H:%M')


            if not work_period_activity or not work_mix_activity:
                continue

            if work_period_activity[0][1] != "entrance":
                #remark_row[11] += " brak wejscia na magazyn"
                remark_row_dict["notes"] += " brak wejscia na magazyn"
                PRINT = True

            if work_period_activity[-1][1] != "exit" and (work_period_activity[-1][0]).day != (CSV_DIAPAZONE[1] - timedelta(hours=8)).day:
                #remark_row[11] += " brak wyjscia z magazynu"
                remark_row_dict["notes"] += " brak wyjscia z magazynu"
                PRINT = True


            if len(work_period_activity) >= 4:
                total_breacfest = get_total_breakfast(work_period_activity).total_seconds() // 60
                breacfest_limit_out = analize_breakfast(datetime_start_shift, datetime_finish_shift, total_breacfest)
                if breacfest_limit_out:
                    #remark_row[7] = breacfest_limit_out
                    remark_row_dict["breakfest_total_time"] = breacfest_limit_out
                    PRINT = True
            else:
                #remark_row[7] = ""
                remark_row_dict["breakfest_total_time"] = ""
            #remark_row[10] = str(datetime_finish_shift - datetime_start_shift - get_total_breakfast(work_period_activity))
            #remark_row_dict["work_total_time"] = str(datetime_finish_shift - datetime_start_shift - get_total_breakfast(work_period_activity))

            remark_row_dict["work_total_time"] = get_work_total_hours(
                datetime_start_shift = datetime_start_shift, 
                datetime_finish_shift = datetime_finish_shift, 
                datetime_start_actualy = actual_start_time,
                datetime_finish_actualy = actual_finish_time,
                breacfest_total = get_total_breakfast(work_period_activity)
            )
            if PRINT:
                #RAPORT_LIST.append(remark_row)
                RAPORT_LIST.append([
                    remark_row_dict["date_start_shift"],
                    remark_row_dict["hour_start_shif"],
                    remark_row_dict["hour_finish_shif"],
                    remark_row_dict["hour_start_work"],
                    remark_row_dict["hour_finish_work"],
                    remark_row_dict["department"],
                    remark_row_dict["full_name"],
                    remark_row_dict["card_id"],
                    remark_row_dict["firm"],
                    remark_row_dict["position"],
                    remark_row_dict["boss"],
                    remark_row_dict["breakfest_total_time"],
                    remark_row_dict["being_late"],
                    remark_row_dict["early_exit"],
                    remark_row_dict["work_total_time"],
                    remark_row_dict["notes"]
                ])

    return RAPORT_LIST


def generate_report(data_to_report):
    title_list = [
        "Data",
        "palnowadana godzina rozpoczęcia",
        "planowana godzina zakończenia",
        "rzeczywista  godzina rozpoczęcia",
        "rzeczywista  godzina zakończenia",
        "Dział",
        "Nazwisko i imię",
        "Nr karty",
        "Firma",
        "Stanowisko",
        "Przełożony",
        "Czas poza magazynem",
        "Spóźnienie",
        "Wcześniejsze wyjście",
        "Rzeczywisty czas pracy",
        "Uwagi"
    ]

    # Sort the data by the first column (assuming it's the date)
    sorted_data = sorted(data_to_report, key=lambda x: x[0])

    # Create a new Excel workbook and add a worksheet
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Report Kolowrotki")

    # Write the title row
    for col_index, title in enumerate(title_list):
        sheet.write(0, col_index, title)

    # Write the data rows
    for row_index, row_data in enumerate(sorted_data, start=1):
        for col_index, cell_data in enumerate(row_data):
            sheet.write(row_index, col_index, cell_data)

    # Save the Excel workbook to a file
    workbook.save("Kolowrotki.xls")



def core(constant_dict):

    global WHITE_LIST_PATH 
    global SCHEDULE_PATH 
    global SCHEDULE_MONTH
    global ACTIVITY_FILE_PATH
    global REFUSAL_FILE_PATH
    global WORKER_DICT_ACTIVITY


    WHITE_LIST_PATH = constant_dict["white_list_path"]
    SCHEDULE_PATH = constant_dict["schedule_path"]
    SCHEDULE_MONTH = constant_dict["schedule_month"]
    ACTIVITY_FILE_PATH = constant_dict["activity_file_path"]
    REFUSAL_FILE_PATH = constant_dict["refusal_file_path"]

    
    generate_white_list()
    generate_workers_schedule_and_information()
    generate_worker_activity()
    data_to_raport = analize_schedule()
    if data_to_raport:
        generate_report(data_to_raport)

